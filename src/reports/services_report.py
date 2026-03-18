from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from src.db.connection import get_connection


def fetch_services():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.id, c.full_name, s.description, s.service_date, s.amount, s.status
        FROM services s
        JOIN clients c ON c.id = s.client_id
        ORDER BY s.id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def export_services_excel(output_dir="exports"):
    rows = fetch_services()
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"

    headers = ["ID", "Cliente", "Descripción", "Fecha", "Monto", "Estado"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r[0],
            r[1] or "",
            r[2] or "",
            str(r[3]) if r[3] is not None else "",
            float(r[4]) if r[4] is not None else 0.0,
            r[5] or ""
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_servicios_{ts}.xlsx"
    wb.save(path)
    return path


def export_services_pdf(output_dir="exports"):
    rows = fetch_services()
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    ts_title = datetime.now().strftime("%d/%m/%Y %H:%M")
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_servicios_{ts_file}.pdf"

    doc = SimpleDocTemplate(
        path,
        pagesize=LETTER,
        leftMargin=50,
        rightMargin=50,
        topMargin=50,
        bottomMargin=50
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Servicios / Pagos", styles["Title"]))
    story.append(Paragraph(f"Generado: {ts_title}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["ID", "Cliente", "Descripción", "Fecha", "Monto", "Estado"]]
    for r in rows:
        data.append([
            str(r[0]),
            r[1] or "",
            r[2] or "",
            str(r[3]) if r[3] is not None else "",
            f"{float(r[4]):.2f}" if r[4] is not None else "0.00",
            r[5] or ""
        ])

    table = Table(data, colWidths=[40, 130, 180, 80, 60, 70])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    story.append(table)
    doc.build(story)
    return path