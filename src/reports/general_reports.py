from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from src.db.connection import get_connection


def ensure_output_dir(output_dir="exports"):
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    return output_dir


def export_clients_excel(output_dir="exports"):
    output_dir = ensure_output_dir(output_dir)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, full_name, phone, email, address
        FROM clients
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["ID", "Nombre", "Teléfono", "Correo", "Dirección"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r[0],
            r[1] or "",
            r[2] or "",
            r[3] or "",
            r[4] or ""
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_clientes_{ts}.xlsx"
    wb.save(path)
    return path


def export_employees_excel(output_dir="exports"):
    output_dir = ensure_output_dir(output_dir)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, full_name, dui, phone, position, salary, active
        FROM employees
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    headers = ["ID", "Nombre", "DUI", "Teléfono", "Cargo", "Salario", "Activo"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r[0],
            r[1] or "",
            r[2] or "",
            r[3] or "",
            r[4] or "",
            float(r[5]) if r[5] is not None else 0.0,
            "Sí" if r[6] else "No"
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_empleados_{ts}.xlsx"
    wb.save(path)
    return path


def export_inventory_excel(output_dir="exports"):
    output_dir = ensure_output_dir(output_dir)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, product_name, unit, quantity, min_stock, cost, notes
        FROM inventory
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["ID", "Producto", "Unidad", "Cantidad", "Stock mínimo", "Costo", "Notas", "Estado"]
    ws.append(headers)

    for r in rows:
        qty = r[3] if r[3] is not None else 0
        min_s = r[4] if r[4] is not None else 0
        estado = "BAJO" if qty <= min_s else "OK"

        ws.append([
            r[0],
            r[1] or "",
            r[2] or "",
            qty,
            min_s,
            float(r[5]) if r[5] is not None else 0.0,
            r[6] or "",
            estado
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_inventario_{ts}.xlsx"
    wb.save(path)
    return path


def export_clients_pdf(output_dir="exports"):
    output_dir = ensure_output_dir(output_dir)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, full_name, phone, email, address
        FROM clients
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    ts_title = datetime.now().strftime("%d/%m/%Y %H:%M")
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_clientes_{ts_file}.pdf"

    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Clientes", styles["Title"]))
    story.append(Paragraph(f"Generado: {ts_title}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["ID", "Nombre", "Teléfono", "Correo", "Dirección"]]
    for r in rows:
        data.append([
            str(r[0]),
            r[1] or "",
            r[2] or "",
            r[3] or "",
            r[4] or ""
        ])

    table = Table(data, colWidths=[35, 120, 90, 120, 150])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    story.append(table)
    doc.build(story)
    return path


def export_employees_pdf(output_dir="exports"):
    output_dir = ensure_output_dir(output_dir)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, full_name, dui, phone, position, salary, active
        FROM employees
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    ts_title = datetime.now().strftime("%d/%m/%Y %H:%M")
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_empleados_{ts_file}.pdf"

    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Empleados", styles["Title"]))
    story.append(Paragraph(f"Generado: {ts_title}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["ID", "Nombre", "DUI", "Teléfono", "Cargo", "Salario", "Activo"]]
    for r in rows:
        data.append([
            str(r[0]),
            r[1] or "",
            r[2] or "",
            r[3] or "",
            r[4] or "",
            f"{float(r[5]):.2f}" if r[5] is not None else "0.00",
            "Sí" if r[6] else "No"
        ])

    table = Table(data, colWidths=[30, 120, 70, 80, 90, 55, 45])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    story.append(table)
    doc.build(story)
    return path


def export_inventory_pdf(output_dir="exports"):
    output_dir = ensure_output_dir(output_dir)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, product_name, unit, quantity, min_stock, cost, notes
        FROM inventory
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    ts_title = datetime.now().strftime("%d/%m/%Y %H:%M")
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{output_dir}/reporte_inventario_{ts_file}.pdf"

    doc = SimpleDocTemplate(path, pagesize=LETTER, leftMargin=35, rightMargin=35, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Inventario", styles["Title"]))
    story.append(Paragraph(f"Generado: {ts_title}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["ID", "Producto", "Unidad", "Cantidad", "Mínimo", "Costo", "Estado"]]
    for r in rows:
        qty = r[3] if r[3] is not None else 0
        min_s = r[4] if r[4] is not None else 0
        estado = "BAJO" if qty <= min_s else "OK"

        data.append([
            str(r[0]),
            r[1] or "",
            r[2] or "",
            str(qty),
            str(min_s),
            f"{float(r[5]):.2f}" if r[5] is not None else "0.00",
            estado
        ])

    table = Table(data, colWidths=[30, 150, 70, 60, 60, 60, 60])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    story.append(table)
    doc.build(story)
    return path